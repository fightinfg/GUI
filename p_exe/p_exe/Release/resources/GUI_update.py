from qpt.run_wrapper import wrapper
wrapper()
import tkinter
import tkinter.messagebox # 导入提示框
import tkinter.simpledialog # 导入对话框
import xlrd

import re # 引入正则模块
import os
from tkinter import filedialog
import pandas as pd
from main4 import *
import sys
def get_resource_path(relative_path):
    if getattr(sys,"frozen",False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path,relative_path)
# 导入图片
IMAGES_PATH_bg= get_resource_path(os.path.join("resources2","背景.png"))  # 定义背景路径
IMAGES_PATH_dt= get_resource_path(os.path.join("resources2","标准答案94-68.png"))  # 定义标准答案路径
#IMAGES_PATH_dk= get_resource_path(os.path.join("resources2","关键词94-68.png"))  # 定义关键词路径
IMAGES_PATH_ds= get_resource_path(os.path.join("resources2","学生答案94-68.png"))  # 定义学生答案路径
IMAGES_PATH_js= get_resource_path(os.path.join("resources2","计算评分94-68.png"))   # 定义学生答案路径
IMAGES_PATH_s= get_resource_path(os.path.join("resources2","后退95-68.png"))   # 定义学生答案路径
IMAGES_PATH_a= get_resource_path(os.path.join("resources2","前进95-68.png"))   # 定义学生答案路径
IMAGES_PATH_save= get_resource_path(os.path.join("resources2","保持结果.png"))   # 定义学生答案路径
IMAGES_PATH_score= get_resource_path(os.path.join("resources2","每题分数94-68.png"))   # 定义学生答案路径
#IMAGES_PATH_k= get_resource_path(os.path.join("resources2","权重系数94-68.png"))   # 定义学生答案路径
# IMAGES_PATH_bg= "resources"+os.sep + "背景.png"  # 定义背景路径
# IMAGES_PATH_dt= "resources"+os.sep + "标准答案94-68.png"  # 定义标准答案路径
# IMAGES_PATH_dk= "resources"+os.sep + "关键词94-68.png"  # 定义关键词路径
# IMAGES_PATH_ds= "resources"+os.sep + "学生答案94-68.png"  # 定义学生答案路径
# IMAGES_PATH_js= "resources"+os.sep + "计算评分94-68.png"  # 定义学生答案路径
# IMAGES_PATH_s= "resources"+os.sep + "后退95-68.png"  # 定义学生答案路径
# IMAGES_PATH_a= "resources"+os.sep + "前进95-68.png"  # 定义学生答案路径
# IMAGES_PATH_save= "resources"+os.sep + "保持结果.png"  # 定义学生答案路径
# IMAGES_PATH_score= "resources"+os.sep + "每题分数94-68.png"  # 定义学生答案路径
# IMAGES_PATH_k= "resources"+os.sep + "权重系数94-68.png"  # 定义学生答案路径
class MainForm:
    def __init__(self):  # 构造方法里面进行窗体控制
        #print(openpyxl.__version__)
        self.root = tkinter.Tk()  # 创建一个窗体
        self.root.title("GUI")  # 设置标题
        # 定义窗体logo
        # self.root.iconbitmap(LOGO_PATH)  # 设置logo资源
        self.root.geometry("1042x734")  # 设置初始化大小
        #self.root.geometry("983x688")  # 设置初始化大小
        # self.root.maxsize(1500, 1024)  # 设置最大的尺寸
        self.photo_bg = tkinter.PhotoImage(file=IMAGES_PATH_bg)
        label_bg = tkinter.Label(self.root, image=self.photo_bg)  # 图片的标签
        label_bg.place(x=0, y=0)  # 背景图片标签的显示

        self.ss1 = 1
        #self.ss2 = 1
        self.ss3 = 1
        self.ss4 = 1
        #self.ss5 = 1


        # 设置标准答案文件导入
        # 设置button按钮接受功能
        self.photo_dt = tkinter.PhotoImage(file=IMAGES_PATH_dt)
        self.button_import_1 = tkinter.Button(self.root, image=self.photo_dt)
        self.button_import_1.bind("<Button-1>", lambda event: self.open_file_1(event))
        self.button_import_1.place(x=22,y=142)
        # 设置entry 但不显示
        self.entry_filename_1 = tkinter.Entry(self.root, width=30, font=("宋体", 10, 'bold'))

        '''
        # 设置关键词文件导入
        # 设置button按钮接受功能
        self.photo_dk = tkinter.PhotoImage(file=IMAGES_PATH_dk)
        self.button_import_2 = tkinter.Button(self.root, image=self.photo_dk)
        self.button_import_2.bind("<Button-1>", lambda event: self.open_file_2(event))
        self.button_import_2.place(x=178, y=340)
        # 设置entry 但不显示
        self.entry_filename_2 = tkinter.Entry(self.root, width=30, font=("宋体", 10, 'bold'))
        '''


        # 设置学生答案文件导入
        # 设置button按钮接受功能
        self.photo_ds = tkinter.PhotoImage(file=IMAGES_PATH_ds)
        self.button_import_3 = tkinter.Button(self.root, image=self.photo_ds)
        self.button_import_3.bind("<Button-1>", lambda event: self.open_file_3(event))
        self.button_import_3.place(x=22, y=260)
        # 设置entry 但不显示
        self.entry_filename_3 = tkinter.Entry(self.root, width=30, font=("宋体", 10, 'bold'))

        # 设置总分文件导入
        # 设置button按钮接受功能
        self.photo_score = tkinter.PhotoImage(file=IMAGES_PATH_score)
        self.button_import_4 = tkinter.Button(self.root, image=self.photo_score)
        self.button_import_4.bind("<Button-1>", lambda event: self.open_file_4(event))
        self.button_import_4.place(x=22, y=380)
        # 设置entry 但不显示
        self.entry_filename_4 = tkinter.Entry(self.root, width=30, font=("宋体", 10, 'bold'))

        '''
        # 设置系数文件导入
        # 设置button按钮接受功能
        self.photo_k = tkinter.PhotoImage(file=IMAGES_PATH_k)
        self.button_import_5 = tkinter.Button(self.root, image=self.photo_k)
        self.button_import_5.bind("<Button-1>", lambda event: self.open_file_5(event))
        self.button_import_5.place(x=30, y=340)
        # 设置entry 但不显示
        self.entry_filename_5 = tkinter.Entry(self.root, width=30, font=("宋体", 10, 'bold'))
        '''


        # 设置计算按钮
        self.photo_js = tkinter.PhotoImage(file=IMAGES_PATH_js)
        self.button_js = tkinter.Button(self.root, image=self.photo_js)
        self.button_js.bind("<Button-1>", lambda event: self.calculate(event))
        self.button_js.place(x=26, y=575)

        # 查看后面一题结果
        self.photo_s = tkinter.PhotoImage(file=IMAGES_PATH_s)
        self.button_s = tkinter.Button(self.root, image=self.photo_s)
        self.button_s.bind("<Button-1>", lambda event: self.retreat(event))
        self.button_s.place(x=513, y=675)

        # 查看前面一题结果
        self.photo_a = tkinter.PhotoImage(file=IMAGES_PATH_a)
        self.button_a = tkinter.Button(self.root, image=self.photo_a)
        self.button_a.bind("<Button-1>", lambda event: self.advance(event))
        self.button_a.place(x=624, y=676)

        # 保存结果
        self.photo_save = tkinter.PhotoImage(file=IMAGES_PATH_save)
        self.button_save = tkinter.Button(self.root, image=self.photo_save)
        self.button_save.bind("<Button-1>", lambda event: self.SAVE(event))
        self.button_save.place(x=736, y=675)


        self.root.mainloop()


    # 打开标准答案文件函数
    def open_file_1(self,event):
        self.entry_filename_1.delete(0, 'end')
        filename = filedialog.askopenfilename(title='打开xls文件', filetypes=[('xls', '*.xls')])
        self.entry_filename_1.insert('insert', filename)
        self.ss1 = self.entry_filename_1.get()  # 用get提取entry中的内容
        print(type(self.ss1))

    '''
        # 打开关键词文件函数
        def open_file_2(self,event):
        self.entry_filename_2.delete(0, 'end')
        filename = filedialog.askopenfilename(title='打开xlsx文件', filetypes=[('xlsx', '*.xlsx')])
        self.entry_filename_2.insert('insert', filename)
        self.ss2 = self.entry_filename_2.get()  # 用get提取entry中的内容
        print(self.ss2)
    '''


    # 打开学生答案文件函数
    def open_file_3(self,event):
        self.entry_filename_3.delete(0, 'end')
        filename = filedialog.askopenfilename(title='打开xls文件', filetypes=[('xls', '*.xls')])
        self.entry_filename_3.insert('insert', filename)
        self.ss3 = self.entry_filename_3.get()  # 用get提取entry中的内容
        print(self.ss3)

    # 打开总分文件函数
    def open_file_4(self, event):
        self.entry_filename_4.delete(0, 'end')
        filename = filedialog.askopenfilename(title='打开xls文件', filetypes=[('xls', '*.xls')])
        self.entry_filename_4.insert('insert', filename)
        self.ss4 = self.entry_filename_4.get()  # 用get提取entry中的内容
        print(self.ss4)

    '''
        def open_file_5(self, event):
        self.entry_filename_5.delete(0, 'end')
        filename = filedialog.askopenfilename(title='打开xlsx文件', filetypes=[('xlsx', '*.xlsx')])
        self.entry_filename_5.insert('insert', filename)
        self.ss5 = self.entry_filename_5.get()  # 用get提取entry中的内容
        print(self.ss5)
    '''
    # 打开系数文件函数


    # 设置主要计算函数
    def calculate(self,event):

        if self.ss1 == 1:  # 设置预警，防止未导入标准答案
            import tkinter.messagebox  # 导入提示框
            tkinter.messagebox.showerror('警告', '标准答案没有导入！')
        #elif self.ss2 == 1:  # 设置预警，防止未导入关键词
        #    import tkinter.messagebox  # 导入提示框
        #    tkinter.messagebox.showerror('警告', '关键词没有导入！')
        elif self.ss3 == 1:  # 设置预警，防止未导入学生答案
            import tkinter.messagebox  # 导入提示框
            tkinter.messagebox.showerror('警告', '学生答案没有导入！')
        elif self.ss4 == 1:  # 设置预警，防止未导入总分文件
            import tkinter.messagebox  # 导入提示框
            tkinter.messagebox.showerror('警告', '总分文件没有导入！')
        #elif self.ss5 == 1:  # 设置预警，防止未导入得分系数文件
        #    import tkinter.messagebox  # 导入提示框
        #    tkinter.messagebox.showerror('警告', '得分系数文件没有导入！')
        else:
            #data_dt = pd.read_excel(self.ss1,engine='openpyxl')
            data_dt = pd.read_excel(self.ss1, engine='xlrd')
            # data_dk = pd.read_excel(self.ss2, engine='openpyxl')
            #data_ds = pd.read_excel(self.ss3, engine='openpyxl')
            data_ds = pd.read_excel(self.ss3, engine='xlrd')
            #data_score = pd.read_excel(self.ss4, engine='openpyxl')
            data_score = pd.read_excel(self.ss4, engine='xlrd')
            # data_k = pd.read_excel(self.ss5, engine='openpyxl')

            list_s1 = data_dt.iloc[:, 1:].values
            list_s2 = data_ds.iloc[:, 1:].values
            # list_key = data_dk.iloc[:, 1:].values
            list_score = data_score.iloc[:, 1:].values
            # list_k_key = data_k.iloc[:, 1].values
            # list_k_yu = data_k.iloc[:, 2].values
            # print(type(float(list_k_key[1])))
            self.xx = len(list_s2[:,0]) # 题目
            self.yy=len(list_s2[0,:]) # 学生
            self.list_Last=[]
            for i in range(0,self.xx):
                '''
                if float(list_k_key[i]) + float(list_k_yu[i]) != 1:
                    import tkinter.messagebox  # 导入提示框
                    tkinter.messagebox.showerror('警告', '第%s题关键词系数和语义系数和不等于1')
                else:
                '''

                List=[]
                # print(i+1,str(list_key[i]))
                for j in range(0,self.yy):
                    s1 = str(list_s1[i,0])
                    s2 = str(list_s2[i,j])
                    # key_words = str(list_key[i])
                    # keyWords = key_words.split(' ')
                    #Result = Review(s1, s2, keyWords)
                    print("s1", s1)
                    print("s2", s2)
                    Result = Review(s1, s2,list_score[i])
                    # print("标准答案否定词个数：", Result["count2"])
                    # print("学生答案否定词个数：", Result["count1"])
                    print("第",i+1,"题，","第",j+1,"人")
                    print("关键词相似度：", Result["Sa1"])
                    print("语义相似度：", Result["Sa2"])
                    print("句长相似度：", Result["Sa4"])
                    '''
                    print("关键词相似度：", round(Result["Sa1"]*10))
                    print("语义相似度：", round(Result["Sa2"]*10))
                    print("句长相似度：", round(Result["Sa4"]*10))
                    '''

                    # print("核心语义相似度：", Result["Sa3"])
                    # print("最终得分：", Result["S"])
                    # print(type(list_score[i]))
                    #key_score = int(round(Result["Sa1"]*float(list_score[i])))
                    #yu_score = int(round(Result["Sa2"]*float(list_score[i])))
                    #last_score = int(round(key_score*list_k_key[i] + yu_score*list_k_yu[i]))
                    last_score=Result["S"]
                    LIST= last_score
                    List.append(LIST)
                self.list_Last.append(List)
            print(self.list_Last)
            # 选择哪个问题表
            self.num = 1
            self.table = self.list_Last[self.num-1]
            import tkinter
            from tkinter import ttk
            self.frame = tkinter.Frame(self.root)

            self.frame.place(height=400,x=320, y=150)
            self.scrollBar = ttk.Scrollbar(self.frame,orient=tkinter.VERTICAL)

            self.tree = ttk.Treeview(self.frame,yscrollcommand=self.scrollBar.set)  # #创建表格对象

            # self.xscroll = ttk.Scrollbar(self.frame, orient=tkinter.HORIZONTAL, command=self.tree.xview)


            # cols=( "第%s题" % str(self.num ),"关键词得分", "语义得分", "最终得分")
            if self.xx > 6:
                self.x1 = 6
                import tkinter.messagebox  # 导入提示框
                tkinter.messagebox.showerror('warn!', 'Only parts of the table can be displayed. \nClick the button of Save, click Result.excel, view the results!')
            else:
                self.x1 = self.xx
            cols = ["Score"]
            for i in range(0,self.x1):
                cols.append("Problem %s" %int(i+1))

            cols = tuple(cols)
            self.tree.column("#0", width=0)
            self.tree["columns"] = cols  # #定义列
            for i in range(0,self.x1+1):
                    self.tree.column("%s" %i, width=90)
                    self.tree.heading("%s" %i,text="%s" %cols[i])
            # self.tree.column("#0", width=0)
            # self.tree["columns"] = cols  # #定义列
            # self.tree.column("第%s题" % str(self.num ), width=100)
            # self.tree.column("关键词得分", width=100)  # #设置列
            # self.tree.column("语义得分", width=100)
            # self.tree.column("最终得分", width=100)
            # self.tree.heading("第%s题" % str(self.num ), text="第%s题" % str(self.num ))  # #设置显示的表头名
            # self.tree.heading("关键词得分", text="关键词得分")# #设置显示的表头名
            # self.tree.heading("语义得分", text="语义得分")
            # self.tree.heading("最终得分", text="最终得分")

            # self.ybar = tkinter.ttk.Scrollbar(self.root, command=self.tree.yview)  # 竖直滚动条
            # self.tree.configure(yscrollcommand=self.ybar.set)
            # self.ybar.place(x=750, y=72)

            # x = len(self.table)
            self.score=np.array(self.list_Last)
            kk = []
            for i in range(0,self.yy):
                # kk.clear
                kk = []
                kk=["Student %s" % str(i+1)]
                for j in range(0,self.x1):
                    # print(self.score[j,i])
                    kk.append(str(self.score[j,i].tolist()))
                # kk = tuple(kk)
                self.tree.insert('','end',values=kk)  # #给第0行添加数据，索引值可重复
                # kk = list(kk)
            self.scrollBar.pack(side=tkinter.RIGHT, fill=tkinter.BOTH)
            # scrollBar.place(x=825, y=72)
            self.scrollBar.config(command=self.tree.yview)

            # self.tree.configure(xscrollcommand=self.xscroll.set)
            # self.xscroll.pack(side=tkinter.BOTTOM, fill=tkinter.X)

            # self.tree.place(x=422,y=72)
            self.tree.pack(side=tkinter.LEFT, fill=tkinter.BOTH)



    # 显示后面一张表
    def retreat(self,event):

        if self.num != 1: # 判断是否为第一张表
            self.num=self.num-1
            self.tree.delete(*self.tree.get_children())
            self.table = self.list_Last[self.num - 1]
            import tkinter
            from tkinter import ttk
            self.frame = tkinter.Frame(self.root)
            self.frame.place(height=400,x=320, y=150)
            self.scrollBar = ttk.Scrollbar(self.frame, orient=tkinter.VERTICAL)
            self.tree = ttk.Treeview(self.frame, yscrollcommand=self.scrollBar.set)  # #创建表格对象
            if self.xx > 6:
                self.x1 = 6
                import tkinter.messagebox  # 导入提示框
                tkinter.messagebox.showerror('warn!',
                                             'Only parts of the table can be displayed. \nClick the button of Save, click Result.excel, view the results!')
            else:
                self.x1 = self.xx
            cols = ["Score"]
            for i in range(0, self.x1):
                cols.append("Problem %s" % int(i + 1))

            cols = tuple(cols)
            self.tree.column("#0", width=0)
            self.tree["columns"] = cols  # #定义列
            for i in range(0, self.x1 + 1):
                self.tree.column("%s" % i, width=95)
                self.tree.heading("%s" % i, text="%s" % cols[i])
            self.score = np.array(self.list_Last)
            kk = []
            for i in range(0, self.yy):
                kk = []
                kk = ["Student %s" % str(i + 1)]
                for j in range(0, self.x1):
                    kk.append(str(self.score[j, i].tolist()))
                self.tree.insert('', 'end', values=kk)  # #给第0行添加数据，索引值可重复
            self.scrollBar.pack(side=tkinter.RIGHT, fill=tkinter.BOTH)
            self.scrollBar.config(command=self.tree.yview)
            self.tree.pack(side=tkinter.LEFT, fill=tkinter.BOTH)

        else:
            import tkinter.messagebox  # 导入提示框
            tkinter.messagebox.showerror('warn!','This is the first table！')

    # 显示前面一张表
    def advance(self,event):

        if self.num != self.xx:  # 判断是否为最后一张表
            self.num=self.num+1
            self.tree.delete(*self.tree.get_children())
            self.table = self.list_Last[self.num - 1]
            import tkinter
            from tkinter import ttk

            self.frame = tkinter.Frame(self.root)

            self.frame.place(height=400,x=320, y=150)
            self.scrollBar = ttk.Scrollbar(self.frame, orient=tkinter.VERTICAL)

            self.tree = ttk.Treeview(self.frame, yscrollcommand=self.scrollBar.set)  # #创建表格对象

            cols = ("Analyze", "The lowest score", "The highest score", "Degree of difficulty",
                    "Coefficient of differentiation")

            self.tree.column("#0", width=0)
            self.tree["columns"] = cols  # #定义列
            self.tree.column("Analyze", width=102)
            self.tree.column("The lowest score", width=119)  # #设置列
            self.tree.column("The highest score", width=119)
            self.tree.column("Degree of difficulty", width=129)
            self.tree.column("Coefficient of differentiation", width=196)
            self.tree.heading("Analyze", text="Analyze")  # #设置显示的表头名
            self.tree.heading("The lowest score", text="The lowest score")  # #设置显示的表头名
            self.tree.heading("The highest score", text="The highest score")
            self.tree.heading("Degree of difficulty", text="Degree of difficulty")
            self.tree.heading("Coefficient of differentiation", text="Coefficient of differentiation")

            import math
            #data_score = pd.read_excel(self.ss4, engine='openpyxl')
            data_score = pd.read_excel(self.ss4, engine='xlrd')
            list_score = data_score.iloc[:, 1:].values
            k_NAN = []
            MAX_score = []
            MIN_score = []
            Separate_score = []
            for i in range(0, self.xx):  # 题目
                aver_score = 0
                max_score = max(self.list_Last[i])
                min_score = min(self.list_Last[i])
                last_score = self.list_Last[i]
                last_score.sort()
                print(last_score)
                num_middle = math.floor(self.yy / 2)
                low_score = 0
                heigh_score = 0
                for j in range(0, self.yy):  # 学生
                    aver_score = aver_score + self.list_Last[i][j]
                    if j <= num_middle - 1:
                        low_score = low_score + last_score[j]
                    else:
                        heigh_score = heigh_score + last_score[j]
                low_score = low_score / num_middle
                heigh_score = heigh_score / (self.yy - num_middle)
                separate = 2 * (heigh_score - low_score) / list_score[i]  # 单题区分系数
                aver_score = aver_score / self.yy  # 每题均分
                k_nan = 1 - aver_score / list_score[i]
                k_NAN.append(float(k_nan))  # 难度系数
                MAX_score.append(max_score)  # 最高得分
                MIN_score.append(min_score)  # 最低得分
                Separate_score.append(float(separate))  # 区分系数



            for i in range(0, self.xx):
                ana = ("Problem %s" % int(i + 1), float(MIN_score[i]), float(MAX_score[i]), k_NAN[i], Separate_score[i])
                self.tree.insert('', 'end', values=ana)
            self.scrollBar.pack(side=tkinter.RIGHT, fill=tkinter.BOTH)
            self.scrollBar.config(command=self.tree.yview)
            self.tree.pack(side=tkinter.LEFT, fill=tkinter.BOTH)

        else:
            import tkinter.messagebox  # 导入提示框
            tkinter.messagebox.showerror('warn!','This is the last table！')

    # 保存结果
    def SAVE(self,event):
        from openpyxl.workbook import Workbook
        from openpyxl.writer.excel import ExcelWriter
        import numpy as np
        wb = Workbook()

        ws = wb.active
        ws.title = u"Score"
        line3 = []
        for i in range(0,self.yy+1): # 学生
            line2=[]
            if i == 0:
                line2=["Score"]
            else:
                line2=["Student %s" %int(i)]
            for j in range(0,self.xx): # 题目
                if i == 0:
                    line1 = "Problem %s" %int(j+1)
                else:
                    line1 = str(self.list_Last[j-1][i-1])
                line2.append(line1)
            line3.append(line2)

        for line4 in line3:
            ws.append(line4)
        import math
        wd = wb.create_sheet("Analyze" )
        wd.title = u"Analyze"
        #data_score = pd.read_excel(self.ss4, engine='openpyxl')
        data_score = pd.read_excel(self.ss4, engine='xlrd')
        list_score = data_score.iloc[:, 1:].values
        k_NAN=[]
        MAX_score = []
        MIN_score = []
        Average_score = []
        Var_score = []
        Separate_score = []
        for i in range(0,self.xx): # 题目
            aver_score = 0
            max_score = max(self.list_Last[i])
            min_score = min(self.list_Last[i])
            average_score = np.mean(np.array(self.list_Last[i]),dtype=np.float64)
            # average_score1 = sum(self.list_Last[i])/len(self.list_Last[i])
            var_score = np.var(np.array(self.list_Last[i]),dtype=np.float64)

            last_score = self.list_Last[i]
            last_score.sort()
            print(last_score)
            num_middle = math.floor(self.yy/2)
            low_score = 0
            heigh_score =0
            for j in range(0,self.yy): # 学生
                aver_score = aver_score + self.list_Last[i][j]
                if j <= num_middle-1:
                    low_score = low_score + last_score[j]
                else:
                    heigh_score = heigh_score +last_score[j]
            low_score = low_score/num_middle
            heigh_score = heigh_score/(self.yy-num_middle)
            separate = 2*(heigh_score-low_score)/list_score[i] # 单题区分系数
            aver_score = aver_score/self.yy # 每题均分
            k_nan = 1-aver_score/list_score[i]
            k_NAN.append(float(k_nan)) # 难度系数
            MAX_score.append(max_score) # 最高得分
            MIN_score.append(min_score) # 最低得分
            Average_score.append(average_score) # 平均得分
            Var_score.append(var_score) # 方差
            Separate_score.append(float(separate)) # 区分系数

        analyze = [["Problem","The lowest score","The highest score","Average score","Variance","Degree of difficulty","Coefficient of differentiation"]]
        for i in range(0,self.xx):
            ana = ["Problem %s" %int(i+1),float(MIN_score[i]),float(MAX_score[i]),float(Average_score[i]),float(Var_score[i]),k_NAN[i],Separate_score[i]]
            print(ana)
            analyze.append(ana)
        for ANA in analyze:
            wd.append(ANA)




        wb.save("Result.xls")



        # wb = Workbook()
        #
        # ws = wb.active
        #
        # for i in range(1,self.xx+1):
        #
        #     if i != 1:
        #         ws = wb.create_sheet("第%s题" %str(i))
        #     ws.title = u"第%s题" % str(i)
        #
        #     self.table = self.list_Last[i-1]
        #     line1=[["关键词得分", "语义得分", "最终得分"]]
        #     for line in self.table:
        #         line1.append(line)
        #     for line2 in line1:
        #         ws.append(line2)
        #     wb.save("结果.xlsx")









MainForm()  # 实例化窗体


